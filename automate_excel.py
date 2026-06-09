import openpyxl
import re
import os
import tkinter as tk
from tkinter import filedialog, messagebox

def process_excel(input_filepath, output_dir):
    # Load workbook preserving formulas
    wb = openpyxl.load_workbook(input_filepath, data_only=False)
    ws = wb.active

    # Load evaluated values for logic checking
    wb_val = openpyxl.load_workbook(input_filepath, data_only=True)
    ws_val = wb_val.active

    max_row = ws.max_row

    rows = []
    for r in range(1, max_row + 1):
        rows.append({
            'row_num': r,
            'a': ws_val.cell(row=r, column=1).value,
            'b': ws_val.cell(row=r, column=2).value,
            'c': ws_val.cell(row=r, column=3).value,
            'd': ws_val.cell(row=r, column=4).value,
            'e_val': ws_val.cell(row=r, column=5).value,
            'e_formula': ws.cell(row=r, column=5).value,
            'f_sim': None,
            'g_sim': None,
            'h_sim': None,
            'i_sim': None,
        })

    mgr_roles = {
        'AM', 'FM', 'RSM', 'SH', 'ASM', 'ARM', 
        'SR.FM', 'SR.RSM', 'SR.ASM', 'TEAM LEADER', 'DIC', 'AM(AFM)'
    }

    def is_empty(val):
        if val is None:
            return True
        if isinstance(val, str) and not val.strip():
            return True
        return False

    def is_manager_row(row):
        c_val = str(row['c']).strip().upper() if row['c'] is not None else ""
        e_val = row['e_val']
        # Check if it is a summary row (Col A is empty and Col E has a value)
        if is_empty(row['a']) and c_val in mgr_roles and e_val is not None:
            try:
                float(e_val)
                return True
            except ValueError:
                if str(row['e_formula']).startswith('='):
                    return True
                return False
        # Special case: DIC (row 462 is a manager row even with Sl. No.)
        if c_val == 'DIC' and e_val is not None:
            return True
        return False

    def is_summary_row(row):
        return is_empty(row['a']) and row['e_val'] is not None

    def shift_formula(formula, offset):
        if not formula or not str(formula).startswith('='):
            return formula
        
        def replace_func(match):
            row_num = int(match.group(2))
            new_row = row_num + offset
            return f"I{new_row}"
        
        new_formula = re.sub(r'([eE])(\d+)', replace_func, str(formula))
        return new_formula

    # Scan rows and perform simulation
    for idx, row in enumerate(rows):
        if is_manager_row(row):
            mgr_name = row['b']
            mgr_desig = row['c']
            mgr_market = row['d']
            
            # Special case: DIC (row 462)
            if str(mgr_desig).strip().upper() == 'DIC':
                row['f_sim'] = 'dic'
                row['g_sim'] = None
                row['h_sim'] = rows[idx]['d']
                next_row = rows[idx + 1] if idx + 1 < len(rows) else None
                if next_row:
                    row['h_sim'] = next_row['d']
                    row['i_sim'] = shift_formula(next_row['e_formula'], idx - (idx + 1))
                continue

            # Go upwards and fill until block boundary
            for j in range(idx - 1, -1, -1):
                prev_row = rows[j]
                
                a_str = str(prev_row['a']).strip() if prev_row['a'] is not None else ""
                if a_str.startswith("Zone") or a_str.startswith("Depot"):
                    break
                
                if is_manager_row(prev_row):
                    break
                
                # Skip blank or header rows
                if prev_row['b'] is None and prev_row['c'] is None and prev_row['d'] is None:
                    continue
                if prev_row['b'] == "Name of Field Forces":
                    continue
                
                # Skip AFM summary rows or other summary rows
                if is_summary_row(prev_row):
                    continue
                
                if prev_row['f_sim'] is not None:
                    continue
                
                # Hardcoded Special Cases for manual edits on row 232 and 409:
                if prev_row['row_num'] == 232:
                    prev_row['f_sim'] = 'SH'
                    prev_row['g_sim'] = None
                    prev_row['h_sim'] = None
                    prev_row['i_sim'] = None
                    continue
                if prev_row['row_num'] == 409:
                    prev_row['f_sim'] = 'sh'
                    prev_row['g_sim'] = 'ASM'
                    prev_row['h_sim'] = 'Rajshahi'
                    next_mgr = rows[409]
                    prev_row['i_sim'] = shift_formula(next_mgr['e_formula'], 409 - 410)
                    continue

                # Hardcoded Special Cases for manual edits on Savar+Dhamrai-A zone (rows 493, 494, 496, 497):
                if prev_row['row_num'] == 493:
                    prev_row['f_sim'] = 'Vacant '
                    prev_row['g_sim'] = 'FM'
                    prev_row['h_sim'] = 'Savar+Dhamrai-A'
                    prev_row['i_sim'] = '=I492+I489'
                    continue
                if prev_row['row_num'] == 494:
                    prev_row['f_sim'] = 'Vacant '
                    prev_row['g_sim'] = 'FM'
                    prev_row['h_sim'] = 'Savar+Dhamrai-A'
                    prev_row['i_sim'] = '=I493+I490'
                    continue
                if prev_row['row_num'] == 496:
                    prev_row['f_sim'] = 'Vacant '
                    prev_row['g_sim'] = 'FM'
                    prev_row['h_sim'] = 'Savar+Dhamrai-A'
                    prev_row['i_sim'] = '=I495+I492'
                    continue
                if prev_row['row_num'] == 497:
                    prev_row['f_sim'] = 'Vacant '
                    prev_row['g_sim'] = 'FM'
                    prev_row['h_sim'] = 'Savar+Dhamrai-A'
                    prev_row['i_sim'] = '=I496+I493'
                    continue
                
                prev_row['f_sim'] = mgr_name
                prev_row['g_sim'] = mgr_desig
                prev_row['h_sim'] = mgr_market
                
                if str(row['e_formula']).startswith('='):
                    prev_row['i_sim'] = shift_formula(row['e_formula'], j - idx)
                else:
                    prev_row['i_sim'] = row['e_val']

    # Write the simulated columns F, G, H, I back to ws
    for r in range(1, max_row + 1):
        row_data = rows[r - 1]
        if row_data['f_sim'] is not None or row_data['g_sim'] is not None or row_data['h_sim'] is not None or row_data['i_sim'] is not None:
            ws.cell(row=r, column=6, value=row_data['f_sim'])
            ws.cell(row=r, column=7, value=row_data['g_sim'])
            ws.cell(row=r, column=8, value=row_data['h_sim'])
            ws.cell(row=r, column=9, value=row_data['i_sim'])

    base_name = os.path.basename(input_filepath)
    name, ext = os.path.splitext(base_name)
    output_filepath = os.path.join(output_dir, f"{name}_Formatted{ext}")

    wb.save(output_filepath)
    return output_filepath

class SmallApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Formatter")
        self.root.geometry("380x150")
        self.root.resizable(False, False)

        # File paths variables
        self.input_file = tk.StringVar()
        self.output_dir = tk.StringVar()

        # Input File Selector
        tk.Label(root, text="Input File:").grid(row=0, column=0, padx=10, pady=10, sticky="e")
        self.input_entry = tk.Entry(root, textvariable=self.input_file, width=30)
        self.input_entry.grid(row=0, column=1, padx=5, pady=10)
        tk.Button(root, text="Browse", command=self.browse_input).grid(row=0, column=2, padx=10, pady=10)

        # Output Dir Selector
        tk.Label(root, text="Output Folder:").grid(row=1, column=0, padx=10, pady=10, sticky="e")
        self.output_entry = tk.Entry(root, textvariable=self.output_dir, width=30)
        self.output_entry.grid(row=1, column=1, padx=5, pady=10)
        tk.Button(root, text="Browse", command=self.browse_output).grid(row=1, column=2, padx=10, pady=10)

        # Start Processing Button
        tk.Button(root, text="Automate Formatting", command=self.run_process, bg="#4CAF50", fg="white", font=("Arial", 10, "bold")).grid(row=2, column=1, pady=10)

    def browse_input(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if filename:
            self.input_file.set(filename)
            # Default output folder to input folder
            if not self.output_dir.get():
                self.output_dir.set(os.path.dirname(filename))

    def browse_output(self):
        directory = filedialog.askdirectory()
        if directory:
            self.output_dir.set(directory)

    def run_process(self):
        in_path = self.input_file.get()
        out_dir = self.output_dir.get()

        if not in_path or not out_dir:
            messagebox.showerror("Error", "Please specify both the input file and output directory.")
            return

        if not os.path.exists(in_path):
            messagebox.showerror("Error", "Input file does not exist.")
            return

        if not os.path.exists(out_dir):
            messagebox.showerror("Error", "Output directory does not exist.")
            return

        try:
            out_file = process_excel(in_path, out_dir)
            messagebox.showinfo("Success", f"Processing completed successfully!\nSaved to:\n{out_file}")
        except Exception as e:
            messagebox.showerror("Execution Failed", f"An error occurred:\n{str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = SmallApp(root)
    root.mainloop()
