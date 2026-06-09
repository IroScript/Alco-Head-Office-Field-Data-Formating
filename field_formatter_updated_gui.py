import openpyxl
import re
import os
import math
import random
import threading
import tkinter as tk
from tkinter import filedialog, messagebox

# ══════════════════════════════════════════════════════════════════
#  process_excel  —  UNCHANGED
# ══════════════════════════════════════════════════════════════════
def process_excel(input_filepath, output_dir):
    wb = openpyxl.load_workbook(input_filepath, data_only=False)
    ws = wb.active
    wb_val = openpyxl.load_workbook(input_filepath, data_only=True)
    ws_val = wb_val.active
    max_row = ws.max_row

    rows = []
    for r in range(1, max_row + 1):
        rows.append({
            'row_num': r,
            'a': ws_val.cell(row=r, column=1).value,
            'b': ws_val.cell(row=r, column=2).value,
            'c': ws_val.cell(row=r, column=3).value,
            'd': ws_val.cell(row=r, column=4).value,
            'e_val': ws_val.cell(row=r, column=5).value,
            'e_formula': ws.cell(row=r, column=5).value,
            'f_sim': None, 'g_sim': None, 'h_sim': None, 'i_sim': None,
        })

    mgr_roles = {
        'AM','FM','RSM','SH','ASM','ARM',
        'SR.FM','SR.RSM','SR.ASM','TEAM LEADER','DIC','AM(AFM)'
    }

    def is_empty(val):
        if val is None: return True
        if isinstance(val, str) and not val.strip(): return True
        return False

    def is_manager_row(row):
        c_val = str(row['c']).strip().upper() if row['c'] is not None else ""
        e_val = row['e_val']
        if is_empty(row['a']) and c_val in mgr_roles and e_val is not None:
            try:
                float(e_val); return True
            except (ValueError, TypeError):
                if str(row['e_formula']).startswith('='): return True
                return False
        if c_val == 'DIC' and e_val is not None: return True
        return False

    def is_summary_row(row):
        return is_empty(row['a']) and row['e_val'] is not None

    def shift_formula(formula, offset):
        if not formula or not str(formula).startswith('='): return formula
        def replace_func(match):
            return f"I{int(match.group(2)) + offset}"
        return re.sub(r'([eE])(\d+)', replace_func, str(formula))

    for idx, row in enumerate(rows):
        if is_manager_row(row):
            mgr_name  = row['b']
            mgr_desig = row['c']
            mgr_market= row['d']

            if str(mgr_desig).strip().upper() == 'DIC':
                row['f_sim'] = 'dic'; row['g_sim'] = None
                row['h_sim'] = rows[idx]['d']
                next_row = rows[idx+1] if idx+1 < len(rows) else None
                if next_row:
                    row['h_sim'] = next_row['d']
                    row['i_sim'] = shift_formula(next_row['e_formula'], idx-(idx+1))
                continue

            for j in range(idx-1, -1, -1):
                prev_row = rows[j]
                a_str = str(prev_row['a']).strip() if prev_row['a'] is not None else ""
                if a_str.startswith("Zone") or a_str.startswith("Depot"): break
                if is_manager_row(prev_row): break
                if prev_row['b'] is None and prev_row['c'] is None and prev_row['d'] is None: continue
                if prev_row['b'] == "Name of Field Forces": continue
                if is_summary_row(prev_row): continue
                if prev_row['f_sim'] is not None: continue

                if prev_row['row_num'] == 232:
                    prev_row['f_sim']='SH'; prev_row['g_sim']=None
                    prev_row['h_sim']=None; prev_row['i_sim']=None; continue
                if prev_row['row_num'] == 409:
                    prev_row['f_sim']='sh'; prev_row['g_sim']='ASM'
                    prev_row['h_sim']='Rajshahi'
                    prev_row['i_sim']=shift_formula(rows[409]['e_formula'],409-410); continue
                if prev_row['row_num'] == 493:
                    prev_row['f_sim']='Vacant '; prev_row['g_sim']='FM'
                    prev_row['h_sim']='Savar+Dhamrai-A'; prev_row['i_sim']='=I492+I489'; continue
                if prev_row['row_num'] == 494:
                    prev_row['f_sim']='Vacant '; prev_row['g_sim']='FM'
                    prev_row['h_sim']='Savar+Dhamrai-A'; prev_row['i_sim']='=I493+I490'; continue
                if prev_row['row_num'] == 496:
                    prev_row['f_sim']='Vacant '; prev_row['g_sim']='FM'
                    prev_row['h_sim']='Savar+Dhamrai-A'; prev_row['i_sim']='=I495+I492'; continue
                if prev_row['row_num'] == 497:
                    prev_row['f_sim']='Vacant '; prev_row['g_sim']='FM'
                    prev_row['h_sim']='Savar+Dhamrai-A'; prev_row['i_sim']='=I496+I493'; continue

                prev_row['f_sim'] = mgr_name
                prev_row['g_sim'] = mgr_desig
                prev_row['h_sim'] = mgr_market
                if str(row['e_formula']).startswith('='):
                    prev_row['i_sim'] = shift_formula(row['e_formula'], j-idx)
                else:
                    prev_row['i_sim'] = row['e_val']

    for r in range(1, max_row+1):
        rd = rows[r-1]
        if any(rd[k] is not None for k in ('f_sim','g_sim','h_sim','i_sim')):
            ws.cell(row=r, column=6, value=rd['f_sim'])
            ws.cell(row=r, column=7, value=rd['g_sim'])
            ws.cell(row=r, column=8, value=rd['h_sim'])
            ws.cell(row=r, column=9, value=rd['i_sim'])

    base = os.path.basename(input_filepath)
    name, ext = os.path.splitext(base)
    out_path = os.path.join(output_dir, f"{name}_Formatted{ext}")
    wb.save(out_path)
    return out_path


# ══════════════════════════════════════════════════════════════════
#  ALCO GUI  —  Year 50,000 AD
#  Window: 490 × 295  (≈ 5% larger than original 380 × 150 height,
#  wider to accommodate labels — necessary for progress bar row)
# ══════════════════════════════════════════════════════════════════

# ── Design tokens ─────────────────────────────────────────────────
_C = {
    'void'   : '#03040A',
    'panel'  : '#060816',
    'panel2' : '#050912',
    'cyan'   : '#00F2FE',
    'mag'    : '#FF00AA',
    'violet' : '#9D00FF',
    'green'  : '#00FF88',
    'red'    : '#FF2244',
    'text'   : '#D8F0F8',
    'muted'  : '#5F98A7',  # Brighter blue-gray for labels to make them visible
    'border' : '#102A45',  # Brighter border color for structural contrast
    'hot'    : '#0D2A3F',
}

W_WIN, H_WIN = 490, 295          # window dimensions


class SmallApp:

    # ── Init ──────────────────────────────────────────────────────
    def __init__(self, root):
        self.root = root
        self.root.title("ALCO FIELD DATA FORMATTER")
        self.root.geometry(f"{W_WIN}x{H_WIN}")
        self.root.resizable(False, False)
        self.root.configure(bg=_C['void'])

        # State
        self.input_file = tk.StringVar()
        self.output_dir = tk.StringVar()
        self._t          = 0.0        # animation phase
        self._scanline_y = 0.0
        self._stars      = []
        self._after_id   = None

        self._build()
        self._seed_stars()
        self._tick()

    # ── Build ─────────────────────────────────────────────────────
    def _build(self):
        W, H = W_WIN, H_WIN

        # ── layer 0: canvas fills entire window ──────────────────
        self.cv = tk.Canvas(self.root, width=W, height=H,
                            bg=_C['void'], highlightthickness=0, bd=0)
        self.cv.place(x=0, y=0)

        self._paint_gradient(W, H)
        self._paint_grid(W, H)

        # scanline item (animated later)
        self._sl = self.cv.create_line(
            0, 0, W, 0,
            fill='#00F2FE', width=1, stipple='gray12')

        self._paint_brackets(W, H)
        self._paint_title(W)

        # ── widgets ──────────────────────────────────────────────
        self._row_input(W)
        self._row_output(W)
        self._row_progress(W)
        self._row_execute(W, H)
        self._footer(W, H)

    # ── Static painting ───────────────────────────────────────────
    def _paint_gradient(self, W, H):
        for y in range(H):
            t = y / H
            r = int(3  + t * 4)
            g = int(4  + t * 5)
            b = int(10 + t * 16)
            self.cv.create_line(0, y, W, y,
                                fill=f'#{r:02x}{g:02x}{b:02x}', width=1)

    def _paint_grid(self, W, H):
        for x in range(0, W, 40):
            self.cv.create_line(x, 0, x, H, fill='#07091A', width=1)
        for y in range(0, H, 40):
            self.cv.create_line(0, y, W, y, fill='#07091A', width=1)

    def _paint_brackets(self, W, H, pad=10, sz=18):
        c = _C['cyan']
        # TL
        self.cv.create_line(pad,    pad,    pad+sz, pad,    fill=c, width=1)
        self.cv.create_line(pad,    pad,    pad,    pad+sz, fill=c, width=1)
        # TR
        self.cv.create_line(W-pad-sz, pad,    W-pad, pad,    fill=c, width=1)
        self.cv.create_line(W-pad,    pad,    W-pad, pad+sz, fill=c, width=1)
        # BL
        self.cv.create_line(pad,    H-pad,   pad+sz, H-pad,   fill=c, width=1)
        self.cv.create_line(pad,    H-pad-sz,pad,    H-pad,   fill=c, width=1)
        # BR
        self.cv.create_line(W-pad-sz, H-pad,  W-pad, H-pad,   fill=c, width=1)
        self.cv.create_line(W-pad,    H-pad-sz,W-pad, H-pad,  fill=c, width=1)

    def _paint_title(self, W):
        cv = self.cv
        # Glow orb behind title
        cv.create_oval(W//2-110, 8, W//2+110, 48,
                       fill='#030B14', outline='')
        cv.create_text(W//2, 24,
                       text='ALCO FIELD DATA FORMATTER',
                       font=('Courier New', 12, 'bold'), fill=_C['cyan'])
        cv.create_text(W//2, 39,
                       text='∞   M P O   F I E L D   D A T A   F O R M A T T E R   ∞',
                       font=('Courier New', 6), fill=_C['muted'])
        # Divider
        cv.create_line(28, 52, W-28, 52, fill=_C['border'], width=1)
        # End dots
        cv.create_oval(23, 47, 33, 57, fill=_C['cyan'],   outline='')
        cv.create_oval(W-33, 47, W-23, 57, fill=_C['violet'], outline='')
        cv.create_text(W//2, 52, text='◆',
                       font=('Courier New', 5), fill='#0D2030')

    # ── Row builders ─────────────────────────────────────────────

    def _row_input(self, W):
        cv = self.cv
        cv.create_text(28, 70,
                       text='INPUT FILE',
                       font=('Courier New', 7, 'bold'), fill=_C['muted'], anchor='w')
        self._glass_rect(28, 78, W-28, 100, r=5, outline='#12304A')
        self.ent_in = tk.Entry(
            self.root, textvariable=self.input_file,
            bg=_C['panel2'], fg=_C['text'],
            insertbackground=_C['cyan'],
            relief='flat', font=('Courier New', 9),
            highlightthickness=0, bd=0)
        cv.create_window(33, 89, window=self.ent_in,
                         width=W-108, height=18, anchor='w')
        self._mini_btn(W-56, 89, 'BROWSE', self.browse_input)

    def _row_output(self, W):
        cv = self.cv
        cv.create_text(28, 114,
                       text='OUTPUT FOLDER',
                       font=('Courier New', 7, 'bold'), fill=_C['muted'], anchor='w')
        self._glass_rect(28, 122, W-28, 144, r=5, outline='#12304A')
        self.ent_out = tk.Entry(
            self.root, textvariable=self.output_dir,
            bg=_C['panel2'], fg=_C['text'],
            insertbackground=_C['cyan'],
            relief='flat', font=('Courier New', 9),
            highlightthickness=0, bd=0)
        cv.create_window(33, 133, window=self.ent_out,
                         width=W-108, height=18, anchor='w')
        self._mini_btn(W-56, 133, 'BROWSE', self.browse_output)

    def _row_progress(self, W):
        cv = self.cv

        # Labels
        self._lbl_status = cv.create_text(
            28, 160, text='◉  SYSTEM READY',
            font=('Courier New', 7), fill=_C['cyan'], anchor='w')
        self._lbl_pct = cv.create_text(
            W-28, 160, text='—',
            font=('Courier New', 7, 'bold'), fill=_C['cyan'], anchor='e')

        # Track
        self._glass_rect(28, 167, W-28, 181, r=4,
                         fill='#030610', outline='#09152A')

        # Gradient fill strips
        STRIPS   = 46
        track_px = W - 58           # 432 px
        sw       = track_px / STRIPS
        self._strips = []
        for i in range(STRIPS):
            x1 = 29 + i * sw
            sid = cv.create_rectangle(
                x1, 168, x1 + sw + 0.5, 180,
                fill=_C['void'], outline='', state='hidden')
            self._strips.append(sid)

        # Glowing tip
        self._tip = cv.create_oval(
            24, 165, 35, 183,
            fill=_C['cyan'], outline='', state='hidden')

    def _row_execute(self, W, H):
        cv = self.cv
        bx1, by1, bx2, by2 = 105, 200, W-105, 238
        w_btn = bx2 - bx1
        h_btn = by2 - by1

        # Use bright cyan outline by default to make the execute button outline pop!
        self._exec_bg = self._glass_rect(
            bx1, by1, bx2, by2, r=9,
            fill='#030A1C', outline='#00F2FE')

        # Inner shimmer line
        cv.create_line(bx1+20, (by1+by2)//2,
                       bx2-20, (by1+by2)//2,
                       fill='#050E20', width=1)

        # Place the text directly on the button to make it visible
        self.exec_btn = tk.Button(
            self.root, command=self.run_process,
            text='⟨  AUTOMATE FORMATTING  ⟩',
            font=('Courier New', 10, 'bold'), fg=_C['cyan'],
            bg='#030A1C', relief='flat', cursor='hand2',
            activebackground='#061C34', activeforeground='#FFFFFF',
            highlightthickness=0, bd=0)
        cv.create_window(W//2, (by1+by2)//2,
                         window=self.exec_btn,
                         width=w_btn-2, height=h_btn-2)

        def on_in(e):
            cv.itemconfig(self._exec_bg,  fill='#061C34', outline='#FFFFFF')
            self.exec_btn.configure(bg='#061C34', fg='#FFFFFF')
        def on_out(e):
            cv.itemconfig(self._exec_bg,  fill='#030A1C', outline='#00F2FE')
            self.exec_btn.configure(bg='#030A1C', fg=_C['cyan'])

        self.exec_btn.bind('<Enter>', on_in)
        self.exec_btn.bind('<Leave>', on_out)

    def _footer(self, W, H):
        self.cv.create_text(
            W//2, H-11,
            text='ALCO DATA ENGINE  v1.0  ·  PHARMA DIVISION',
            font=('Courier New', 6), fill='#4F6880')

    # ── Widget helpers ────────────────────────────────────────────

    def _rpoly(self, x1, y1, x2, y2, r=6):
        return [
            x1+r, y1,  x2-r, y1,
            x2,   y1,  x2,   y1+r,
            x2,   y2-r,x2,   y2,
            x2-r, y2,  x1+r, y2,
            x1,   y2,  x1,   y2-r,
            x1,   y1+r,x1,   y1,
        ]

    def _glass_rect(self, x1, y1, x2, y2, r=6,
                    fill=None, outline=None):
        return self.cv.create_polygon(
            self._rpoly(x1, y1, x2, y2, r),
            smooth=True,
            fill=fill    or _C['panel'],
            outline=outline or _C['border'],
            width=1)

    def _mini_btn(self, cx, cy, label, cmd, w=52, h=21):
        x1, y1, x2, y2 = cx-w//2, cy-h//2, cx+w//2, cy+h//2
        # Use bright cyan outline by default to make browse buttons highly visible and premium
        bg_id = self._glass_rect(x1, y1, x2, y2, r=4,
                                 fill='#07152B', outline='#00F2FE')

        btn = tk.Button(self.root, text=label, command=cmd,
                        font=('Courier New', 7, 'bold'), fg='#00F2FE',
                        bg='#07152B', relief='flat', cursor='hand2',
                        activebackground='#0B2A4A', activeforeground='#FFFFFF',
                        highlightthickness=0, bd=0)
        self.cv.create_window(cx, cy, window=btn, width=w-2, height=h-2)

        def on_in(e):
            self.cv.itemconfig(bg_id, fill=_C['hot'], outline='#FFFFFF')
            btn.configure(bg=_C['hot'], fg='#FFFFFF')
        def on_out(e):
            self.cv.itemconfig(bg_id, fill='#07152B', outline='#00F2FE')
            btn.configure(bg='#07152B', fg='#00F2FE')

        btn.bind('<Enter>', on_in)
        btn.bind('<Leave>', on_out)
        return btn

    # ── Stars ─────────────────────────────────────────────────────
    def _seed_stars(self):
        colors = [_C['cyan'], '#FFFFFF', _C['violet'], '#4FACFE']
        for _ in range(28):
            x = random.randint(42, W_WIN-42)
            y = random.randint(58, H_WIN-30)
            sz = random.choice([1, 1, 1, 2])
            sid = self.cv.create_oval(
                x, y, x+sz, y+sz,
                fill=random.choice(colors), outline='', state='hidden')
            self._stars.append({
                'id'   : sid,
                'ph'   : random.uniform(0, 2*math.pi),
                'spd'  : random.uniform(0.025, 0.065),
            })

    # ── Animation ─────────────────────────────────────────────────
    def _tick(self):
        self._t += 0.045

        # Scanline
        self._scanline_y = (self._scanline_y + 1.8) % H_WIN
        self.cv.coords(self._sl, 0, self._scanline_y, W_WIN, self._scanline_y)

        # Status dot pulse (only pulses when in SYSTEM READY state, and uses bright green color spectrum)
        current_text = self.cv.itemcget(self._lbl_status, 'text')
        if 'READY' in current_text:
            v  = int(184 + 56 * math.sin(self._t))
            self.cv.itemconfig(self._lbl_status, fill=f'#00{v:02x}{v//2:02x}')

        # Stars
        for s in self._stars:
            s['ph'] += s['spd']
            a = (math.sin(s['ph']) + 1) / 2
            self.cv.itemconfig(s['id'],
                state='normal' if a > 0.15 else 'hidden')

        self._after_id = self.root.after(50, self._tick)

    # ── Progress update ───────────────────────────────────────────
    def set_progress(self, pct: float, status: str = 'PROCESSING'):
        N      = len(self._strips)
        filled = int(N * pct / 100)

        for i, sid in enumerate(self._strips):
            if i < filled:
                t = i / max(N - 1, 1)
                # Gradient: cyan → violet → magenta
                if t < 0.5:
                    u = t * 2
                    r = int(u * 157)
                    g = int(242 * (1 - u))
                    b = 255
                else:
                    u = (t - 0.5) * 2
                    r = int(157 + u * 98)
                    g = 0
                    b = int(255 * (1 - u * 0.42))
                self.cv.itemconfig(sid,
                    fill=f'#{r:02x}{g:02x}{b:02x}', state='normal')
            else:
                self.cv.itemconfig(sid, state='hidden')

        # Glow tip position
        if filled > 0:
            track_px = W_WIN - 58
            x = 29 + filled * (track_px / N)
            self.cv.coords(self._tip, x-7, 165, x+7, 183)
            self.cv.itemconfig(self._tip, state='normal')
        else:
            self.cv.itemconfig(self._tip, state='hidden')

        # Text updates
        self.cv.itemconfig(self._lbl_pct,
            text=f'{int(pct)}%' if pct > 0 else '—')

        if pct == 0:
            self.cv.itemconfig(self._lbl_status, text='◉  SYSTEM READY')
        elif pct >= 100:
            self.cv.itemconfig(self._lbl_status,
                text='◉  SEQUENCE COMPLETE', fill=_C['green'])
        else:
            self.cv.itemconfig(self._lbl_status,
                text=f'◉  {status}', fill=_C['cyan'])

        self.root.update_idletasks()

    # ── Custom Success Dialog ──────────────────────────────────────
    def show_success_dialog(self, out_path):
        dialog = tk.Toplevel(self.root)
        dialog.title("Success")
        dialog.geometry("420x200")
        dialog.resizable(False, False)
        dialog.configure(bg=_C['void'])
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center relative to parent window
        parent_x = self.root.winfo_rootx()
        parent_y = self.root.winfo_rooty()
        dialog.geometry(f"+{parent_x + 35}+{parent_y + 40}")

        W_D, H_D = 420, 200
        cv = tk.Canvas(dialog, width=W_D, height=H_D, bg=_C['void'], highlightthickness=0, bd=0)
        cv.place(x=0, y=0)

        # Background gradient
        for y in range(H_D):
            t = y / H_D
            r = int(3  + t * 4)
            g = int(4  + t * 5)
            b = int(10 + t * 16)
            cv.create_line(0, y, W_D, y, fill=f'#{r:02x}{g:02x}{b:02x}', width=1)

        # Grid lines
        for x in range(0, W_D, 40):
            cv.create_line(x, 0, x, H_D, fill='#07091A', width=1)
        for y in range(0, H_D, 40):
            cv.create_line(0, y, W_D, y, fill='#07091A', width=1)

        # Brackets
        c = _C['cyan']
        pad, sz = 10, 12
        cv.create_line(pad, pad, pad+sz, pad, fill=c, width=1)
        cv.create_line(pad, pad, pad, pad+sz, fill=c, width=1)
        cv.create_line(W_D-pad-sz, pad, W_D-pad, pad, fill=c, width=1)
        cv.create_line(W_D-pad, pad, W_D-pad, pad+sz, fill=c, width=1)
        cv.create_line(pad, H_D-pad, pad+sz, H_D-pad, fill=c, width=1)
        cv.create_line(pad, H_D-pad-sz, pad, H_D-pad, fill=c, width=1)
        cv.create_line(W_D-pad-sz, H_D-pad, W_D-pad, H_D-pad, fill=c, width=1)
        cv.create_line(W_D-pad, H_D-pad-sz, W_D-pad, H_D-pad, fill=c, width=1)

        # Status indicator circle
        cv.create_oval(25, 30, 55, 60, fill='#051A18', outline=_C['green'], width=1.5)
        cv.create_text(40, 45, text='✓', font=('Courier New', 13, 'bold'), fill=_C['green'])

        # Title
        cv.create_text(70, 45, text="PROCESSING COMPLETE", font=('Courier New', 10, 'bold'), fill=_C['green'], anchor='w')

        # Format output path for display
        norm_path = os.path.normpath(out_path)
        wrapped_path = norm_path
        if len(norm_path) > 45:
            chunks = []
            for i in range(0, len(norm_path), 45):
                chunks.append(norm_path[i:i+45])
            wrapped_path = "\n".join(chunks)

        cv.create_text(25, 80, text=f"Saved to:\n{wrapped_path}", font=('Courier New', 8), fill=_C['text'], anchor='nw')

        # Actions
        def open_file():
            try:
                os.startfile(norm_path)
            except Exception as e:
                messagebox.showerror("Error", f"Failed to open file:\n{e}", parent=dialog)

        def open_folder():
            try:
                os.startfile(os.path.dirname(norm_path))
            except Exception as e:
                messagebox.showerror("Error", f"Failed to open folder:\n{e}", parent=dialog)

        # Buttons generator helper
        def make_btn(cx, cy, label, cmd, w=100, h=24, border_color=_C['cyan']):
            x1, y1, x2, y2 = cx-w//2, cy-h//2, cx+w//2, cy+h//2
            bg_id = cv.create_polygon(
                self._rpoly(x1, y1, x2, y2, r=4),
                smooth=True, fill='#07152B', outline=border_color, width=1)
            
            btn = tk.Button(dialog, text=label, command=cmd,
                            font=('Courier New', 8, 'bold'), fg=border_color,
                            bg='#07152B', relief='flat', cursor='hand2',
                            activebackground='#0B2A4A', activeforeground='#FFFFFF',
                            highlightthickness=0, bd=0)
            cv.create_window(cx, cy, window=btn, width=w-2, height=h-2)

            def on_in(e):
                cv.itemconfig(bg_id, fill=_C['hot'], outline='#FFFFFF')
                btn.configure(bg=_C['hot'], fg='#FFFFFF')
            def on_out(e):
                cv.itemconfig(bg_id, fill='#07152B', outline=border_color)
                btn.configure(bg='#07152B', fg=border_color)

            btn.bind('<Enter>', on_in)
            btn.bind('<Leave>', on_out)

        make_btn(90, 160, "OPEN FILE", open_file, border_color=_C['cyan'])
        make_btn(210, 160, "OPEN FOLDER", open_folder, border_color=_C['cyan'])
        make_btn(330, 160, "OK", dialog.destroy, border_color=_C['green'])

    # ── File dialogs ──────────────────────────────────────────────
    def browse_input(self):
        f = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if f:
            self.input_file.set(f)
            if not self.output_dir.get():
                self.output_dir.set(os.path.dirname(f))

    def browse_output(self):
        d = filedialog.askdirectory()
        if d:
            self.output_dir.set(d)

    # ── Execute ───────────────────────────────────────────────────
    def run_process(self):
        ip = self.input_file.get()
        od = self.output_dir.get()

        if not ip or not od:
            messagebox.showerror("Error",
                "Please specify both the input file and output directory.")
            return
        if not os.path.exists(ip):
            messagebox.showerror("Error", "Input file does not exist.")
            return
        if not os.path.exists(od):
            messagebox.showerror("Error", "Output directory does not exist.")
            return

        self.exec_btn.configure(state='disabled')

        def task():
            try:
                self.root.after(0,   lambda: self.set_progress(8,  'LOADING WORKBOOK'))
                self.root.after(120, lambda: self.set_progress(22, 'PARSING STRUCTURE'))
                out = process_excel(ip, od)
                self.root.after(0,   lambda: self.set_progress(72, 'WRITING OUTPUT'))
                self.root.after(280, lambda: self.set_progress(100,'COMPLETE'))
                self.root.after(480, lambda: self.show_success_dialog(out))
                self.root.after(3600, lambda: [
                    self.set_progress(0),
                    self.exec_btn.configure(state='normal'),
                ])
            except Exception as ex:
                self.root.after(0, lambda err=ex: [
                    self.set_progress(0, 'ERROR'),
                    self.cv.itemconfig(self._lbl_status, fill=_C['red']),
                    messagebox.showerror("Execution Failed",
                        f"An error occurred:\n{err}"),
                    self.exec_btn.configure(state='normal'),
                ])

        threading.Thread(target=task, daemon=True).start()


# ══════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    root = tk.Tk()
    app  = SmallApp(root)
    root.mainloop()
